import openpyxl
import datetime
from datetime import date, timedelta
from dateutil import parser
otchet = openpyxl.load_workbook('Восеба - май.xlsx')
moscow = openpyxl.load_workbook('Васебо - шаблон.xlsx')
otchet_1 = openpyxl.load_workbook('before_comments.xlsx', data_only = True)
karusel = otchet['Карусель']
karusel_m = moscow['Карусель']
metro = otchet['Метро']
metro_m = moscow['Метро']
perek = otchet['Перекрёсток']
perek_m = moscow['Перекрёсток']


karusel_1 = otchet_1['Карусель']
metro_1 = otchet_1['Метро']
perek_1 = otchet_1['Перекрёсток']

for i in range (2,20):
    print (metro_1.cell(row=i, column=28).value)

for i in range (2,22):
    if karusel_1.cell(row=i, column=25).value==None and karusel_1.cell(row=i, column=24).value<1 and karusel_m.cell(row=i, column=17).value !=None:
        karusel.cell(row=j, column=25, value= karusel_m.cell(row=i,column=17).value)
    elif karusel_1.cell(row=i, column=25).value==None and karusel_1.cell(row=i, column=24).value<1 and karusel_m.cell(row=i, column=17).value ==None:
        karusel.cell(row=j, column=28, value='Ожидается поставка с' + ' ' + str(today))

for i in range (2,20):
    if metro_1.cell(row=i, column=29).value==None and metro_1.cell(row=i, column=28).value<1 and metro_m.cell(row=i, column=21).value !=None:
        metro.cell(row=j, column=29, value= metro_m.cell(row=i,column=21).value)
    elif metro_1.cell(row=i, column=29).value==None and metro_1.cell(row=i, column=28).value<1 and metro_m.cell(row=i, column=21).value ==None:
        metro.cell(row=j, column=29, value='Ожидается поставка с '+ str(today))

for i in range (2,46):
    if perek_1.cell(row=i, column=22).value==None and perek_1.cell(row=i, column=21).value<1 and perek_m.cell(row=i, column=14).value !=None:
        perek.cell(row=j, column=22, value= perek_m.cell(row=i,column=21).value)
    elif perek_1.cell(row=i, column=22).value==None and perek_1.cell(row=i, column=21).value<1 and perek_m.cell(row=i, column=14).value ==None:
        perek.cell(row=j, column=22, value='Ожидается поставка с '+ str(today))


otchet.save('woseba' + str(today) + '.xlsx')
