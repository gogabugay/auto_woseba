import openpyxl
import datetime
from datetime import date, timedelta
from dateutil import parser
otchet = openpyxl.load_workbook('Восеба - 11.05.xlsx')
moscow = openpyxl.load_workbook('Васебо - шаблон.xlsx')
otchet_1 = openpyxl.load_workbook('Восеба - 11.05 (копия).xlsx', data_only=True)


today = date.today()
result = otchet['Отчёт']
result.cell(row = 2, column = 3, value = today)

karusel_1 = otchet_1['Карусель']
metro_1 = otchet_1['Метро']
perek_1 = otchet_1['Перекрёсток']


karusel = otchet['Карусель']
karusel_m = moscow['Карусель']

for b in range (2,22):
    karusel.cell(row=b, column=14, value=karusel_m.cell(row=b, column=6).value.date())

for c in range (2,22):
    for i in range (15,23):
        karusel.cell(row=c, column=i, value=karusel_m.cell(row=c, column=i-8).value)
        karusel_1.cell(row=c, column=i, value=karusel_m.cell(row=c, column=i-8).value)
#print(karusel.cell(row=3, column=24, read_only=True).value)

metro = otchet['Метро']
metro_m = moscow['Метро']

for b in range (2,20):
    metro.cell(row=b, column=14, value=metro_m.cell(row=b, column=6).value.date())

for c in range (2,20):
    for i in range (15,27):
        metro.cell(row=c, column=i, value=metro_m.cell(row=c, column=i-8).value)
        metro_1.cell(row=c, column=i, value=metro_m.cell(row=c, column=i-8).value)



perek = otchet['Перекрёсток']
perek_m = moscow['Перекрёсток']

for b in range (2,46):
    perek.cell(row=b, column=14, value=perek_m.cell(row=b, column=6).value.date())

for c in range (2,46):
    for i in range (15,20):
        perek.cell(row=c, column=i, value=perek_m.cell(row=c, column=i-8).value)
        perek_1.cell(row=c, column=i, value=perek_m.cell(row=c, column=i-8).value)
#otchet.save('before_comments.xlsx')

for i in range (2,22):
    count=0
    count1=0
    for j in range (15,23):
        if karusel_1.cell(row=i, column=j).value == 'х':
            count+=1
            karusel_1.cell(row=i, column=13, value=8-count)
        elif karusel_1.cell(row=i, column=j).value==1:
            count1+=1
            karusel_1.cell(row=i, column=23, value=count1)
for g in range (2,22):
    karusel_1.cell(row=g, column=24, value= karusel_1.cell(row=g, column=23).value / karusel_1.cell(row=g, column=13).value)

for i in range (2,20):
    count=0
    count1=0
    for j in range (15,27):
        if metro_1.cell(row=i, column=j).value == 'х':
            count+=1
            metro_1.cell(row=i, column=13, value=12-count)
        elif metro_1.cell(row=i, column=j).value==1:
            count1+=1
            metro_1.cell(row=i, column=27, value=count1)
for g in range (2,20):
    metro_1.cell(row=g, column=28, value= metro_1.cell(row=g, column=27).value / metro_1.cell(row=g, column=13).value)

for i in range (2,46):
    count=0
    count1=0
    for j in range (15,20):
        if perek_1.cell(row=i, column=j).value == 'х':
            count+=1
            perek_1.cell(row=i, column=13, value=5-count)
        elif perek_1.cell(row=i, column=j).value==1:
            count1+=1
            perek_1.cell(row=i, column=20, value=count1)
for g in range (2,46):
    perek_1.cell(row=g, column=21, value= perek_1.cell(row=g, column=20).value / perek_1.cell(row=g, column=13).value)

for i in range (2,22):
    karusel.cell(row=i, column=25, value= karusel_m.cell(row=i,column=17).value)

for i in range (2,20):
    metro.cell(row=i, column=29, value= metro_m.cell(row=i,column=21).value)

for i in range (2,46):
    perek.cell(row=j, column=22, value= perek_m.cell(row=i,column=14).value)

for i in range (2,22):
    if karusel.cell(row=i, column=25).value==None and karusel_1.cell(row=i, column=24).value<1:
        karusel.cell(row=i, column=28, value='Ожидается поставка с' + ' ' + str(today))
for i in range (2,20):
    if metro.cell(row=i, column=29).value==None and metro_1.cell(row=i, column=28).value<1:
        metro.cell(row=i, column=29, value='Ожидается поставка с '+ str(today))

for i in range (2,46):
    if perek.cell(row=i, column=22).value==None and perek_1.cell(row=i, column=21).value<1:
        perek.cell(row=i, column=22, value='Ожидается поставка с '+ str(today))

otchet_1.save('test.xlsx')
otchet.save('woseba' + str(today) + '.xlsx')
